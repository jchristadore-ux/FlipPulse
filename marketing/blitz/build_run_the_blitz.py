#!/usr/bin/env python3
"""Build **Run_the_Blitz.xlsx** — the one document you open each morning to run
the 90-day blitz without hunting through files.

One tab per campaign week (Week 01 … Week 13), each a top-to-bottom execution
checklist: every scheduled post for that week, in clock order, with its caption,
image/video prompt pointers, CTA, a Status dropdown, a Priority, and Notes. A
"▶ Start Here" tab holds the Founders 100 Club offer, the legend, your editable
links, and the week index.

It is generated from the real campaign so it can never drift:
  • 02_CONTENT_CALENDAR_90D.md — which asset goes in which slot, per day.
  • 03_POST_LIBRARY.md         — the actual caption text (T##, FB##, IG##, FL-*, ST-*).
  • 04_VIDEO_SCRIPTS.md        — video titles + the TikTok/Reels-native idea banks.

The Founders 100 Club launch offer is threaded in from **Day 2 (Jul 14)**; Day 1
(Jul 13) and Week 1 are marked COMPLETE/frozen and left alone. Update the live
"[N]/100 seats left" count as you go and stop once the cap fills.

Regenerate:  python3 marketing/blitz/build_run_the_blitz.py
(Needs openpyxl:  pip install openpyxl)
"""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).parent
CALENDAR = HERE / "02_CONTENT_CALENDAR_90D.md"
LIBRARY = HERE / "03_POST_LIBRARY.md"
SCRIPTS = HERE / "04_VIDEO_SCRIPTS.md"
OUT = HERE / "Run_the_Blitz.xlsx"

YEAR = 2026
FOUNDER_CUTOVER = date(2026, 7, 14)  # Day 2 — offer push begins here
DAY1 = date(2026, 7, 13)             # COMPLETE / frozen — never modify

# ── palette (mirrors the brand kit in 05_GRAPHICS_PROMPTS.md) ─────────────────
MIDNIGHT = "0B0F1A"
INK = "13233A"
GREEN = "00E5A0"
GREEN_SOFT = "0F2A22"
AMBER = "FFC24D"
AMBER_SOFT = "3A3320"
SLATE = "8B93A7"
PAPER = "F5F7FA"
RED = "FF4D5E"
BANNER = "12324F"
DAYBAND = "1B2536"
ZEBRA = "141A28"

FOUNDERS_CTA = ("Founders 100 Club — first 100 join free: $0 to start "
                "(setup + first month waived), then $99/mo locked in for life. go.flippulse.com")
PROOF_CTA = "Watch it decide, live: t.me/flippulse"
CAPTURE_CTA = "The 9 rules it checks, free: [lead magnet link]"
STORY_CTA = "Link sticker → t.me/flippulse"

FOUNDERS_ONELINER = ("🎟 FOUNDERS 100 CLUB — first 100 members join free ($0 to start: "
                     "$150 setup + first month waived), then $99/mo LOCKED IN FOR LIFE + "
                     "Founder status. Post it early & often and update “[N]/100 seats left” "
                     "until the cap fills, then retire it. Assets: 03 §2.0 (FL-X/FL-X2/FL-X3/"
                     "FL-COUNT/FL-FB) · graphic G-3.9 · video V31.")

STANDING_DAILY = ("STANDING EVERY DAY (08_DAILY_OPERATING_PROCEDURE): 8:15 X-am · 9:00 Story · "
                  "12:00 Reel · 12:30 X-mid · 1:00 FB · 6:00 TikTok · 7:30 X-pm  ||  15 X replies AM "
                  "+ 5 PM · reply to ALL comments <12h · forward 1–3 bot alerts to t.me/flippulse & "
                  "grab today’s [shot] · log KPIs PM. During launch: post/refresh the Founders seat count.")

# Clock order so the sheet reads morning → night (col label -> (Platform, time))
SLOT_ORDER = [
    ("X am",  "X · Twitter",     "8:15am"),
    ("Story", "Instagram Story", "9:00am"),
    ("Reel",  "Instagram Reel",  "12:00pm"),
    ("X mid", "X · Twitter",     "12:30pm"),
    ("FB",    "Facebook",        "1:00pm"),
    ("TT",    "TikTok",          "6:00pm"),
    ("X pm",  "X · Twitter",     "7:30pm"),
]

HEADERS = ["Date", "Platform", "Deliverable", "Caption", "Image prompt",
           "Video prompt", "CTA", "Status", "Priority", "Notes"]
COL_W = [16, 17, 20, 60, 26, 30, 40, 12, 10, 34]
STATUS_CHOICES = '"☐ To do,▶ In progress,✅ Done,⏭ Skipped"'


# ─────────────────────────── source parsing ─────────────────────────────────
def _unwrap(t: str) -> str:
    return t.replace("\\n", " ").replace("**", "").replace("`", "").strip()


def load_captions() -> dict[str, str]:
    """T##/FB##/IG## (backtick form), FL-* (bold-label form), ST-* (prose form)."""
    caps: dict[str, str] = {}
    tick = re.compile(r"^- `((?:T|FB|IG)\d+)`(?: \([^)]*\))? `(.*)`\s*$")
    fl = re.compile(r"^- \*\*[^*]+\*\* `(FL-[A-Z0-9]+)`(?: \([^)]*\))?: `(.+)`[^`]*$")
    st = re.compile(r"^- `(ST-[A-Z])`\s+(.*)$")
    for line in LIBRARY.read_text(encoding="utf-8").splitlines():
        m = tick.match(line) or fl.match(line)
        if m:
            caps[m.group(1)] = _unwrap(m.group(2))
            continue
        m = st.match(line)
        if m:
            caps[m.group(1)] = _unwrap(m.group(2))
    return caps


def load_video_meta():
    """V## -> title; X-TT-# and X-RL-# -> one-line idea (from 04 §3/§4)."""
    titles: dict[str, str] = {}
    xtt: dict[str, str] = {}
    xrl: dict[str, str] = {}
    text = SCRIPTS.read_text(encoding="utf-8")
    for m in re.finditer(r'^### (V\d+) — "([^"]+)"', text, re.M):
        titles[m.group(1)] = m.group(2)
    # numbered idea banks
    def _ideas(header: str) -> dict[str, str]:
        out: dict[str, str] = {}
        block = text.split(header, 1)
        if len(block) < 2:
            return out
        body = block[1].split("\n## ", 1)[0].split("\n#### ", 1)[0]
        for m in re.finditer(r"^(\d+)\.\s+`?([^`\n]+?)`?(?:\s+—.*)?$", body, re.M):
            out[m.group(1)] = m.group(2).strip()
        return out
    xtt = _ideas("## §3 — 30 extra TikTok-native ideas")
    xrl = _ideas("## §4 — 30 extra Reels-native ideas")
    return titles, xtt, xrl


WEEK_HDR = re.compile(r"^## WEEK (\d+) \(([^)]*)\)\s*[—-]+\s*(.*)$")


def parse_calendar():
    """Return list of weeks: {num, dates, focus, days:[{label,date,slots:{col:cell}, notes}]}."""
    lines = CALENDAR.read_text(encoding="utf-8").splitlines()
    weeks = []
    i = 0
    while i < len(lines):
        m = WEEK_HDR.match(lines[i])
        if not m:
            i += 1
            continue
        wk = {"num": int(m.group(1)), "dates": m.group(2).strip(),
              "focus": m.group(3).strip(), "days": []}
        i += 1
        # find the table header
        while i < len(lines) and not lines[i].lstrip().startswith("|"):
            i += 1
        if i >= len(lines):
            weeks.append(wk)
            break
        header = [c.strip() for c in lines[i].strip().strip("|").split("|")]
        i += 2  # skip header + separator
        # map calendar columns
        def find(name):
            for idx, h in enumerate(header):
                if h == name or h.startswith(name):
                    return idx
            return None
        col = {"Day": find("Day"), "X am": find("X am"), "X mid": find("X mid"),
               "X pm": find("X pm"), "TT": find("TT"), "Reel": find("Reel"),
               "FB": find("FB"), "Story": find("Story")}
        notes_idx = None
        for idx, h in enumerate(header):
            if h == "Notes" or "Notes" in h or h.startswith("Engage"):
                notes_idx = idx
        while i < len(lines) and lines[i].lstrip().startswith("|"):
            cells = [c.strip() for c in lines[i].strip().strip("|").split("|")]
            i += 1
            if col["Day"] is None or col["Day"] >= len(cells):
                continue
            label = cells[col["Day"]]
            dm = re.search(r"(\d{1,2})/(\d{1,2})", label)
            if not dm:
                continue
            d = date(YEAR, int(dm.group(1)), int(dm.group(2)))
            slots = {}
            for key in ("X am", "X mid", "X pm", "TT", "Reel", "FB", "Story"):
                ci = col[key]
                if ci is not None and ci < len(cells) and cells[ci] not in ("", "—"):
                    slots[key] = cells[ci]
            notes = cells[notes_idx] if (notes_idx is not None and notes_idx < len(cells)) else ""
            wk["days"].append({"label": label, "date": d, "slots": slots, "notes": notes})
        weeks.append(wk)
    return weeks


# ─────────────────────────── row synthesis ──────────────────────────────────
TOKEN_ORDER = [r"TH-PIN", r"FL-[A-Z0-9]+", r"TH\d+", r"C\d+", r"V\d+",
               r"X-TT-\d+", r"X-RL-\d+", r"T\d+", r"FB\d+", r"IG\d+", r"ST-[A-Z]"]


def primary_token(cell: str) -> str:
    for pat in TOKEN_ORDER:
        m = re.search(rf"\b{pat}\b", cell)
        if m:
            return m.group(0)
    return ""


def resolve(platform_col, cell, caps, vtitles, xtt, xrl):
    """Return (caption, image_prompt, video_prompt)."""
    tok = primary_token(cell)
    shot = "[shot]" in cell or "[shots]" in cell
    cap = img = vid = ""

    if tok.startswith("TH"):
        cap = f"🧵 Thread ({tok}) — build the chain in Metricool (+ Add tweet). Full text: 03 §2/§3.2."
    elif tok.startswith("FL-"):
        cap = caps.get(tok, f"Founders 100 Club offer post ({tok}) — 03 §2.0.")
    elif tok in caps and tok[:1] in ("T", "F", "I", "S"):
        cap = caps[tok]
    elif tok.startswith("V"):
        cap = f"🎥 {vtitles.get(tok, tok)} — film per 04 §2 ({tok})."
        ig = re.search(r"IG\d+", cell)
        if ig and ig.group(0) in caps:
            cap += f"  IG caption: {caps[ig.group(0)]}"
    elif tok.startswith("C"):
        cap = f"🖼 Carousel ({tok}) — outline 03 §4.3, build in Canva per 05 §4."
    elif tok.startswith("X-TT-"):
        n = tok.rsplit("-", 1)[1]
        cap = f"🎥 TikTok-native idea #{n}: {xtt.get(n, '(04 §3)')}"
    elif tok.startswith("X-RL-"):
        n = tok.rsplit("-", 1)[1]
        cap = f"🎥 Reels-native idea #{n}: {xrl.get(n, '(04 §4)')}"
    else:
        cap = _unwrap(cell)

    # image prompt
    if shot:
        img = "Fresh <48h Telegram screenshot"
    elif platform_col == "Story":
        img = "Story card — 05 G-4.2"
    elif tok.startswith("C"):
        img = "Carousel — 05 §4 (G-4.1)"
    elif tok.startswith("FL-"):
        img = "Founders card — 05 G-3.9"
    elif platform_col in ("TT", "Reel") and (tok.startswith("V") or tok.startswith("X-")):
        img = "Thumbnail — 05 G-4.4"

    # video prompt
    if platform_col in ("TT", "Reel") and (tok.startswith("V") or tok.startswith("X-")):
        vid = f"Film/edit per 04 ({tok}). Export TikTok + clean (Reel/Short)."

    return cap, img, vid


def cta_for(platform_col, cell, tok, launch_phase):
    if tok.startswith("FL-"):
        return FOUNDERS_CTA
    if tok in ("T45", "T57", "FB10") or "lead magnet" in cell.lower():
        return CAPTURE_CTA
    if tok in ("T42", "T43", "T44") and launch_phase:
        return FOUNDERS_CTA
    if platform_col == "Story":
        return STORY_CTA
    return PROOF_CTA


def priority_for(cell, tok):
    low = cell.lower()
    if tok.startswith("FL-") or "pin" in low or tok in ("TH-PIN", "V10"):
        return "★ High"
    if tok.startswith("TH") or tok.startswith("C") or tok in ("T30", "V30") or "batch" in low:
        return "Med"
    return "Normal"


def notes_for(cell, d):
    flags = []
    low = cell.lower()
    if "pin" in low:
        flags.append("PIN")
    if "[shot]" in cell or "[shots]" in cell:
        flags.append("fresh screenshot")
    if "thread" in low or primary_token(cell).startswith("TH"):
        flags.append("thread")
    if "poll" in low:
        flags.append("poll")
    if "rerun" in low:
        flags.append("rerun → new hook/cut")
    if "batch" in low:
        flags.append("Sunday BATCH")
    if re.search(r"R0\d", cell):
        flags.append("Reddit — post manually + babysit")
    if d == DAY1:
        flags.append("DAY 1 — COMPLETE, do not modify")
    return " · ".join(flags)


# ─────────────────────────── workbook styling ───────────────────────────────
def font(sz=10, bold=False, color=PAPER):
    return Font(name="Calibri", size=sz, bold=bold, color=color)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


TOP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)
thin = Side(style="thin", color="26314A")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_sheet_base(ws):
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(COL_W, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_banner(ws, row, text, hexfill, color=PAPER, size=10, bold=True, height=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(hexfill)
    c.font = font(size, bold, color)
    c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
    if height:
        ws.row_dimensions[row].height = height


# ─────────────────────────── build week tabs ────────────────────────────────
def build_week(ws, wk, caps, vtitles, xtt, xrl):
    style_sheet_base(ws)
    r = 1
    add_banner(ws, r, f"WEEK {wk['num']:02d}  ·  {wk['dates']}  —  {wk['focus']}",
               BANNER, GREEN, size=12, bold=True, height=34); r += 1

    launch_phase = wk["num"] <= 6
    if wk["num"] == 1:
        add_banner(ws, r, "✅ WEEK 1 — COMPLETE / FROZEN. Reference only; do not modify. "
                          "The Founders 100 Club push begins Day 2 (Jul 14).",
                   INK, AMBER, height=28); r += 1
    add_banner(ws, r, FOUNDERS_ONELINER if launch_phase else
               "Founders 100 Club still open? Keep posting it (early & often) until 100 seats fill — "
               "rotate FL-X/FL-X2/FL-X3/FL-COUNT/FL-FB (03 §2.0) and update the seat count.",
               GREEN_SOFT, GREEN, bold=False, height=42 if launch_phase else 28); r += 1
    add_banner(ws, r, STANDING_DAILY, AMBER_SOFT, PAPER, bold=False, height=42); r += 1
    r += 1  # spacer

    # column headers
    hdr_row = r
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.fill = fill(INK); c.font = font(10, True, GREEN); c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[r].height = 18
    r += 1

    dv = DataValidation(type="list", formula1=STATUS_CHOICES, allow_blank=True)
    ws.add_data_validation(dv)

    zebra = False
    for day in wk["days"]:
        d = day["date"]
        frozen = d <= DAY1
        band = f"── {day['label']}  ·  {d.isoformat()} " + ("· DAY 1 COMPLETE ──" if d == DAY1
                else "· complete/past ──" if frozen else "──")
        add_banner(ws, r, band, DAYBAND, SLATE if frozen else PAPER, bold=True, height=16); r += 1

        # Founders daily task from Day 2 within launch weeks
        if launch_phase and d >= FOUNDER_CUTOVER:
            vals = [d.isoformat(), "★ Founders 100 Club",
                    "Post the Club + update [N]/100 seats",
                    "Rotate FL-X/FL-X2/FL-X3/FL-COUNT/FL-FB (03 §2.0). Also drop the seat count in an IG Story.",
                    "Founders card — 05 G-3.9", "Optional: V31 offer video — 04 §2",
                    FOUNDERS_CTA, "☐ To do", "★ High",
                    "Launch offer — early & often until cap fills"]
            _write_row(ws, r, vals, dv, zebra=zebra, accent=GREEN_SOFT); r += 1
            zebra = not zebra

        # engagement / notes for the day (if present)
        if day["notes"]:
            vals = [d.isoformat(), "Engage / focus", _unwrap(day["notes"]),
                    "", "", "", "", "☐ To do", "Normal",
                    "DAY 1 — do not modify" if d == DAY1 else ""]
            _write_row(ws, r, vals, dv, zebra=zebra); r += 1
            zebra = not zebra

        for key, platform, time in SLOT_ORDER:
            if key not in day["slots"]:
                continue
            cell = day["slots"][key]
            tok = primary_token(cell)
            cap, img, vid = resolve(key, cell, caps, vtitles, xtt, xrl)
            cta = cta_for(key, cell, tok, launch_phase and not frozen)
            prio = priority_for(cell, tok)
            note = notes_for(cell, d)
            deliverable = _unwrap(cell)
            vals = [d.isoformat(), f"{platform} · {time}", deliverable, cap, img, vid,
                    cta, "✅ Done" if frozen else "☐ To do", prio, note]
            _write_row(ws, r, vals, dv, zebra=zebra,
                       accent=GREEN_SOFT if tok.startswith("FL-") else None,
                       muted=frozen); r += 1
            zebra = not zebra

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)


def _write_row(ws, r, vals, dv, zebra=False, accent=None, muted=False):
    base = accent or (ZEBRA if zebra else MIDNIGHT)
    for ci, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=ci, value=v)
        c.fill = fill(base)
        color = SLATE if muted else PAPER
        if ci == 9 and isinstance(v, str) and v.startswith("★"):
            color = AMBER
        c.font = font(9.5, bold=(ci == 9 and v.startswith("★")), color=color)
        c.alignment = CENTER if ci in (2, 8, 9) else TOP
        c.border = BORDER
    dv.add(ws.cell(row=r, column=8))
    ws.row_dimensions[r].height = 30


# ─────────────────────────── Start Here tab ─────────────────────────────────
def build_start(ws, weeks):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 46
    for col in "DEFGHIJ":
        ws.column_dimensions[col].width = 14

    def band(row, text, hexfill, color=PAPER, size=11, bold=True, h=None, span=3):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill(hexfill); c.font = font(size, bold, color)
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
        if h:
            ws.row_dimensions[row].height = h

    r = 1
    band(r, "▶ RUN THE BLITZ — your daily execution cockpit", BANNER, GREEN, 14, True, 32); r += 2
    band(r, "How to use: open THIS file every morning → go to this week’s tab → work the rows top to "
            "bottom (they’re in clock order). Set each Status when done. Nothing to hunt for — the "
            "caption, prompts, and CTA are on the row. Full source text lives in 03/04/05 if you want more.",
         INK, PAPER, 10, False, 46); r += 2

    band(r, "🎟 THE OFFER — FOUNDERS 100 CLUB (lead with it, early & often, until 100 seats fill)",
         GREEN_SOFT, GREEN, 12, True, 24); r += 1
    for line in [
        "• The first 100 members join FREE: we waive the $150 setup fee AND the first month — $0 today, then $99/mo, cancel anytime.",
        "• Founder status for LIFE: that $99/mo is grandfathered forever + a permanent Founder badge + first access to everything we ship.",
        "• Real hard cap of 100 seats (Stripe coupon FOUNDING100). Say the remaining count out loud: “[N]/100 seats left”.",
        "• Build posts AROUND the Club — don’t bolt it on the end. Assets: 03 §2.0 (FL-X/FL-X2/FL-X3/FL-COUNT/FL-FB) · graphic 05 G-3.9 · video 04 V31.",
        "• Compliance: price + status only, never returns. The DAY the cap fills, stop advertising it everywhere (11 §7).",
    ]:
        band(r, line, MIDNIGHT, PAPER, 10, False, 26); r += 1
    r += 1

    band(r, "LEGEND", INK, GREEN, 11, True, 20); r += 1
    band(r, "Status: ☐ To do · ▶ In progress · ✅ Done · ⏭ Skipped (dropdown in each Status cell).",
         MIDNIGHT, PAPER, 10, False, 20); r += 1
    band(r, "Priority: ★ High (Founders / pinned / viral candidate — do first) · Med (threads, carousels, "
            "milestones) · Normal.", MIDNIGHT, PAPER, 10, False, 20); r += 1
    band(r, "Posting clock (ET): 8:15 X-am · 9:00 IG Story · 12:00 IG Reel · 12:30 X-mid · 1:00 FB · "
            "6:00 TikTok · 7:30 X-pm. Reddit is NEVER scheduled — post manually, mid-morning.",
         MIDNIGHT, PAPER, 10, False, 20); r += 1
    band(r, "“Frozen”: Week 1 + Day 1 (Jul 13) are COMPLETE — reference only, do not modify.",
         MIDNIGHT, AMBER, 10, False, 20); r += 2

    # editable links table
    band(r, "YOUR LINKS / PLACEHOLDERS (fill these once, reuse everywhere)", INK, GREEN, 11, True, 20); r += 1
    link_hdr = r
    for ci, h in enumerate(["What", "Placeholder in the copy", "Your real URL (edit)"], start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.fill = fill(BANNER); c.font = font(10, True, GREEN); c.alignment = CENTER; c.border = BORDER
    r += 1
    for what, ph in [
        ("Signup / checkout", "go.flippulse.com"),
        ("Live paper feed", "t.me/flippulse"),
        ("Lead magnet (9 Rules)", "[lead magnet link]"),
        ("Link-in-bio", "beacons.ai/flippulse"),
        ("Scheduler", "Metricool → Planning"),
        ("KPI sheet", "kpi_tracker_template.csv → Google Sheets"),
        ("Founders seat count", "Stripe → coupon FOUNDING100 redemptions ([N]/100)"),
    ]:
        for ci, v in enumerate([what, ph, ""], start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = fill(MIDNIGHT if ci < 3 else AMBER_SOFT)
            c.font = font(10, False, PAPER); c.alignment = TOP; c.border = BORDER
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1

    # week index
    band(r, "WEEK INDEX", INK, GREEN, 11, True, 20); r += 1
    for ci, h in enumerate(["Tab", "Dates", "Focus"], start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.fill = fill(BANNER); c.font = font(10, True, GREEN); c.alignment = CENTER; c.border = BORDER
    r += 1
    for wk in weeks:
        for ci, v in enumerate([f"Week {wk['num']:02d}", wk["dates"], wk["focus"]], start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = fill(MIDNIGHT); c.font = font(10, False, PAPER); c.alignment = TOP; c.border = BORDER
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1
    band(r, "Compliance one-liner: “Would I be fine if a regulator, a journalist, and my angriest customer "
            "all read this today?” If not, cut the number, add the disclaimer, or lead with the risk (11 §8). "
            "Short disclaimer: “Not financial advice. Trading involves risk of loss.”",
         AMBER_SOFT, PAPER, 10, False, 40); r += 1


# ─────────────────────────── main ───────────────────────────────────────────
def main() -> None:
    caps = load_captions()
    vtitles, xtt, xrl = load_video_meta()
    weeks = parse_calendar()

    wb = Workbook()
    start = wb.active
    start.title = "▶ Start Here"
    start.sheet_properties.tabColor = GREEN
    build_start(start, weeks)

    for wk in weeks:
        ws = wb.create_sheet(title=f"Week {wk['num']:02d}")
        ws.sheet_properties.tabColor = AMBER if wk["num"] <= 6 else SLATE
        build_week(ws, wk, caps, vtitles, xtt, xrl)

    wb.save(OUT)
    total = sum(len(w["days"]) for w in weeks)
    print(f"Wrote {OUT.name}: {len(weeks)} week tabs + Start Here, {total} campaign-days.")
    print(f"  captions resolved: {len(caps)} · video titles: {len(vtitles)}")


if __name__ == "__main__":
    main()
