#!/usr/bin/env python3
"""
Build FlipPulse_Business_Model.xlsx — a formula-driven operating/budget model.

Edit the blue INPUT cells in the workbook (or the defaults here) and every output
recomputes in Excel. Regenerate with:  python business/build_model.py

Sheets:
  Inputs         — all editable assumptions (pricing, costs, growth, hardware)
  Monthly Model  — 12-month P&L (customers, revenue, costs, capex, net, cumulative)
  Summary        — quarterly + annual rollups and key metrics
  Cost Notes     — researched pricing + sources behind the defaults
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── palette ───────────────────────────────────────────────────────────────────
DARK   = "0B1220"
TEAL   = "35CEE6"
GREEN  = "7EF2A0"
INK    = "1F2A3D"
INPUTF = "FFF4CE"      # editable input cell fill (soft yellow)
HEADF  = "12324F"      # section header fill
WHITE  = "FFFFFF"

def title_font(sz=14, color=WHITE): return Font(bold=True, size=sz, color=color)
def hdr_font():  return Font(bold=True, color=WHITE)
def inp_fill():  return PatternFill("solid", fgColor=INPUTF)
def head_fill(): return PatternFill("solid", fgColor=HEADF)
thin = Side(style="thin", color="D0D7E2")
border = Border(bottom=thin)

USD  = '$#,##0'
USD2 = '$#,##0.00'
PCT  = '0.0%'
NUM  = '#,##0'

# Editable ramp: ACTIVE customers at END of each month (targets: 100 @ m6, 500 @ m12)
ACTIVE_END = [8, 20, 38, 60, 80, 100, 140, 190, 255, 330, 415, 500]

wb = Workbook()

# ══════════════════════════════════════════════════════════════════ INPUTS ════
ws = wb.active
ws.title = "Inputs"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 60

def put(row, label, value=None, *, section=False, fmt=None, note=None, formula=None):
    a = ws.cell(row=row, column=1, value=label)
    if section:
        a.font = hdr_font(); a.fill = head_fill()
        ws.cell(row=row, column=2).fill = head_fill()
        ws.cell(row=row, column=3).fill = head_fill()
        return
    a.font = Font(color="30425F")
    if formula is not None:
        b = ws.cell(row=row, column=2, value=formula)
    elif value is not None:
        b = ws.cell(row=row, column=2, value=value)
        b.fill = inp_fill()                       # editable
    else:
        b = ws.cell(row=row, column=2)
    if fmt:
        b.number_format = fmt
    b.alignment = Alignment(horizontal="right")
    if note:
        ws.cell(row=row, column=3, value=note).font = Font(italic=True, color="6B7890", size=9)

ws.cell(row=1, column=1, value="FlipPulse — Business Model Inputs").font = title_font(15)
ws.cell(row=2, column=1,
        value="Edit the shaded (yellow) cells. Everything else recomputes on the other sheets."
        ).font = Font(italic=True, color="6B7890")

put(4,  "REVENUE", section=True)
put(5,  "Startup fee per new customer", 150, fmt=USD)
put(6,  "Monthly subscription per customer", 99, fmt=USD)
put(7,  "Performance fee (% of customer monthly profit)", 0.0, fmt=PCT,
        note="PLACEHOLDER — temporarily removed / not currently charged. Set e.g. 0.20 to re-enable.")
put(8,  "Include performance fee in headline totals? (1=yes, 0=no)", 0,
        note="Leave 0 while the fee is on hold; needs compliance review before enabling.")

put(10, "PERFORMANCE-FEE ASSUMPTIONS  (illustrative — UNPROVEN)", section=True)
put(11, "Avg funded balance per customer", 2000, fmt=USD,
        note="Your guess of the typical account size.")
put(12, "Avg monthly net return per customer", 0.05, fmt=PCT,
        note="The bot's real edge is unproven (paper mode by default). Adjust freely.")

put(14, "GROWTH", section=True)
put(15, "Monthly churn (% of active customers)", 0.04, fmt=PCT)

put(17, "RECURRING COSTS  (monthly)", section=True)
put(18, "Claude Max", 100, fmt=USD)
put(19, "Railway base (Pro plan)", 20, fmt=USD)
put(20, "Railway usage per active bot", 5, fmt=USD,
        note="Each customer = one always-on bot. ~$3–8/mo of usage each; tune to real bills.")
put(21, "GitHub", 0, fmt=USD, note="Free covers unlimited private repos. Team is $4/user if you add staff.")
put(22, "Misc software / tools (domain, email, etc.)", 30, fmt=USD)

put(24, "PAYMENT PROCESSING  (Stripe)", section=True)
put(25, "Stripe % per charge", 0.029, fmt=PCT)
put(26, "Stripe fixed per charge", 0.30, fmt=USD2)
put(27, "Stripe Billing % (recurring)", 0.007, fmt=PCT)

put(29, "HARDWARE  (capital expense — funded from profits)", section=True)
put(30, "Mac mini (top spec, M4 Pro 48GB / 1TB)", 2000, fmt=USD)
put(31, "Large curved monitor", 1000, fmt=USD)
put(32, "Peripherals / desk setup", 300, fmt=USD)
put(33, "Hardware total", formula="=B30+B31+B32", fmt=USD)
put(34, "Purchase in month (1–12)", 5,
        note="Only sensible once cumulative profit covers it — see the Summary check.")

# input cell refs
R = dict(startup="Inputs!$B$5", sub="Inputs!$B$6", perf="Inputs!$B$7", incl="Inputs!$B$8",
         bal="Inputs!$B$11", ret="Inputs!$B$12", churn="Inputs!$B$15",
         claude="Inputs!$B$18", railbase="Inputs!$B$19", railbot="Inputs!$B$20",
         github="Inputs!$B$21", misc="Inputs!$B$22",
         spct="Inputs!$B$25", sfix="Inputs!$B$26", sbill="Inputs!$B$27",
         hwtot="Inputs!$B$33", hwmon="Inputs!$B$34")

# ═══════════════════════════════════════════════════════════ MONTHLY MODEL ════
m = wb.create_sheet("Monthly Model")
m.sheet_view.showGridLines = False
cols = [
    ("Month", NUM, 8), ("Cust. start", NUM, 11), ("Cust. end", NUM, 11),
    ("Churned", NUM, 10), ("New", NUM, 9),
    ("Setup rev", USD, 12), ("Subscription", USD, 13), ("Perf fee*", USD, 12),
    ("Total revenue", USD, 14),
    ("Claude", USD, 10), ("Railway", USD, 11), ("GitHub+Misc", USD, 12), ("Stripe fees", USD, 12),
    ("Total OpEx", USD, 13), ("Hardware capex", USD, 14),
    ("Net profit", USD, 13), ("Cumulative", USD, 14),
]
m.cell(row=1, column=1, value="FlipPulse — 12-Month Model").font = title_font(14)
for j, (name, fmt, w) in enumerate(cols, start=1):
    c = m.cell(row=2, column=j, value=name)
    c.font = hdr_font(); c.fill = head_fill()
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    m.column_dimensions[get_column_letter(j)].width = w
m.freeze_panes = "A3"

# column letters
A,B,C,D,E,F,G,H,I,J,K,L,M_,N,O,P,Q = (get_column_letter(i) for i in range(1,18))

first = 3
for idx in range(12):
    r = first + idx
    mn = idx + 1
    avg = f"(({B}{r}+{C}{r})/2)"
    m.cell(row=r, column=1, value=mn)                                   # Month
    # Cust start = prior end (0 for month 1)
    m.cell(row=r, column=2, value=(0 if idx == 0 else f"={C}{r-1}"))
    ce = m.cell(row=r, column=3, value=ACTIVE_END[idx]); ce.fill = inp_fill()  # editable ramp
    m.cell(row=r, column=4, value=f"={B}{r}*{R['churn']}")              # churned
    m.cell(row=r, column=5, value=f"={C}{r}-{B}{r}+{D}{r}")             # new
    m.cell(row=r, column=6, value=f"={E}{r}*{R['startup']}")            # setup rev
    m.cell(row=r, column=7, value=f"={avg}*{R['sub']}")                 # subscription
    m.cell(row=r, column=8,                                             # perf fee (illustrative)
           value=f"={avg}*{R['bal']}*{R['ret']}*{R['perf']}*{R['incl']}")
    m.cell(row=r, column=9, value=f"={F}{r}+{G}{r}+{H}{r}")             # total revenue
    m.cell(row=r, column=10, value=f"={R['claude']}")                  # Claude
    m.cell(row=r, column=11, value=f"={R['railbase']}+{avg}*{R['railbot']}")   # Railway
    m.cell(row=r, column=12, value=f"={R['github']}+{R['misc']}")      # GitHub+Misc
    m.cell(row=r, column=13,                                            # Stripe fees
           value=(f"={E}{r}*({R['spct']}*{R['startup']}+{R['sfix']})"
                  f"+{avg}*({R['spct']}*{R['sub']}+{R['sfix']}+{R['sbill']}*{R['sub']})"
                  f"+{H}{r}*({R['spct']}+{R['sbill']})"))
    m.cell(row=r, column=14, value=f"={J}{r}+{K}{r}+{L}{r}+{M_}{r}")    # total opex
    m.cell(row=r, column=15, value=f"=IF({A}{r}={R['hwmon']},{R['hwtot']},0)")  # hardware capex
    m.cell(row=r, column=16, value=f"={I}{r}-{N}{r}-{O}{r}")            # net profit
    m.cell(row=r, column=17,                                            # cumulative
           value=(f"={P}{r}" if idx == 0 else f"={Q}{r-1}+{P}{r}"))
    for j in range(1, 18):
        cell = m.cell(row=r, column=j)
        if j >= 6:
            cell.number_format = USD
        cell.border = border

# totals row
tr = first + 12
m.cell(row=tr, column=1, value="YEAR").font = hdr_font()
for j in (6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16):
    L_ = get_column_letter(j)
    tc = m.cell(row=tr, column=j, value=f"=SUM({L_}{first}:{L_}{first+11})")
    tc.number_format = USD; tc.font = Font(bold=True)
m.cell(row=tr, column=17, value=f"={Q}{first+11}").font = Font(bold=True)
m.cell(row=tr, column=17).number_format = USD
m.cell(row=tr+1, column=1,
       value="* Performance fee is a PLACEHOLDER — currently 0% (not charged). Set Inputs!B7 + B8 to re-enable."
       ).font = Font(italic=True, color="6B7890", size=9)

# ═══════════════════════════════════════════════════════════════ SUMMARY ══════
s = wb.create_sheet("Summary")
s.sheet_view.showGridLines = False
s.column_dimensions["A"].width = 42
for col in "BCDE":
    s.column_dimensions[col].width = 15
s.cell(row=1, column=1, value="FlipPulse — Summary").font = title_font(15)

def srow(r, label, formula=None, fmt=USD, bold=False, section=False):
    a = s.cell(row=r, column=1, value=label)
    if section:
        a.font = hdr_font(); a.fill = head_fill()
        for cc in range(2, 6):
            s.cell(row=r, column=cc).fill = head_fill()
        return
    a.font = Font(bold=bold, color="30425F")
    if formula is not None:
        b = s.cell(row=r, column=2, value=formula); b.number_format = fmt
        b.font = Font(bold=bold)

# Quarterly table
s.cell(row=3, column=1, value="QUARTERLY", ).font = hdr_font()
s.cell(row=3, column=1).fill = head_fill()
for cc, name in enumerate(["", "Q1 (m1-3)", "Q2 (m4-6)", "Q3 (m7-9)", "Q4 (m10-12)"]):
    c = s.cell(row=3, column=cc+1, value=name if cc else "QUARTERLY")
    c.font = hdr_font(); c.fill = head_fill(); c.alignment = Alignment(horizontal="center")
qranges = [("B", "C", "D", "E")]  # placeholder
# rows sum from Monthly Model
def qsum(colletter, q):
    lo = 3 + (q-1)*3
    hi = lo + 2
    return f"=SUM('Monthly Model'!{colletter}{lo}:{colletter}{hi})"
metrics = [("Revenue", "I"), ("Operating costs", "N"), ("Hardware capex", "O"), ("Net profit", "P")]
for i,(lbl,col) in enumerate(metrics):
    r = 4+i
    s.cell(row=r, column=1, value=lbl).font = Font(color="30425F", bold=(lbl=="Net profit"))
    for q in range(1,5):
        cc = s.cell(row=r, column=1+q, value=qsum(col,q))
        cc.number_format = USD; cc.font = Font(bold=(lbl=="Net profit"))

# Annual + key metrics
srow(10, "ANNUAL (Year 1)", section=True)
srow(11, "Total revenue (contracted: setup + subscription)", "=SUM('Monthly Model'!I3:I14)", bold=True)
srow(12, "Total operating costs", "=SUM('Monthly Model'!N3:N14)")
srow(13, "Total hardware capex", "=SUM('Monthly Model'!O3:O14)")
srow(14, "Net profit (Year 1)", "=SUM('Monthly Model'!P3:P14)", bold=True)
srow(15, "Gross margin (rev − opex) / rev",
     "=(SUM('Monthly Model'!I3:I14)-SUM('Monthly Model'!N3:N14))/SUM('Monthly Model'!I3:I14)", fmt=PCT)

srow(17, "KEY METRICS", section=True)
srow(18, "Ending customers (month 12)", "='Monthly Model'!C14", fmt=NUM, bold=True)
srow(19, "Ending MRR (run-rate: customers × sub)", "='Monthly Model'!C14*Inputs!$B$6")
srow(20, "Ending ARR (MRR × 12)", "='Monthly Model'!C14*Inputs!$B$6*12", bold=True)
srow(21, "Avg revenue / customer / month (m12)",
     "='Monthly Model'!I14/(('Monthly Model'!B14+'Monthly Model'!C14)/2)")
srow(22, "Cumulative net profit (end of year)", "='Monthly Model'!Q14", bold=True)

srow(24, "HARDWARE CHECK", section=True)
srow(25, "Cumulative profit BEFORE hardware, at purchase month",
     "=SUMIF('Monthly Model'!$A$3:$A$14,\"<=\"&Inputs!$B$34,'Monthly Model'!$P$3:$P$14)"
     "+SUMIF('Monthly Model'!$A$3:$A$14,\"<=\"&Inputs!$B$34,'Monthly Model'!$O$3:$O$14)")
srow(26, "Hardware total", "=Inputs!$B$33")
srow(27, "Covered by profits?  (should be TRUE)",
     "=IF(B25>=B26,\"YES — funded from profit\",\"NO — push purchase later\")", fmt="General", bold=True)

srow(29, "PERFORMANCE-FEE UPSIDE  (PLACEHOLDER — currently 0%, not charged)", section=True)
srow(30, "Annual performance-fee revenue (0 while on hold)",
     "=(SUMPRODUCT(('Monthly Model'!B3:B14+'Monthly Model'!C3:C14)/2))*Inputs!$B$11*Inputs!$B$12*Inputs!$B$7",
     bold=True)
srow(31, "Total revenue INCL. perf-fee upside",
     "=SUM('Monthly Model'!I3:I14)+B30")
srow(32, "Net profit INCL. perf-fee upside",
     "=SUM('Monthly Model'!P3:P14)+B30", bold=True)
s.cell(row=33, column=1,
       value="Performance fee is temporarily removed (Inputs!B7 = 0), so this reads $0. Set a % on the "
             "Inputs sheet to model it as an upside — unproven, so treat as a ceiling, not a forecast."
       ).font = Font(italic=True, color="6B7890", size=9)

# ═════════════════════════════════════════════════════════════ COST NOTES ═════
n = wb.create_sheet("Cost Notes")
n.sheet_view.showGridLines = False
n.column_dimensions["A"].width = 34
n.column_dimensions["B"].width = 20
n.column_dimensions["C"].width = 74
n.cell(row=1, column=1, value="FlipPulse — Cost Research & Sources (as of mid-2026)").font = title_font(13)
notes = [
    ("Item", "Figure used", "Notes / source"),
    ("Claude Max", "$100 / mo", "Max tier as specified. A $200/mo (higher-usage) tier also exists."),
    ("Railway", "$20 base + ~$5/bot", "Pro plan is $20/mo incl. credits; usage billed per-minute. Each always-on bot is a small container (~$3–8/mo). Hobby is $5/mo incl. $5 usage. (docs.railway.com/pricing)"),
    ("GitHub", "$0", "Free plan includes unlimited private repos. GitHub Team is $4/user/mo if you add collaborators. Copilot is separate ($0–$100/mo)."),
    ("Stripe", "2.9% + $0.30, +0.7%", "Standard US card fee 2.9% + $0.30 per charge; Stripe Billing adds ~0.5–0.7% on recurring. No flat monthly fee on standard. (stripe.com/pricing, stripe.com/billing/pricing)"),
    ("Mac mini (top spec)", "~$2,000", "M4 Pro (14-core) starts ~$1,599 (2026 price rise); 48GB/1TB config lands near $2,000+. (apple.com/mac-mini)"),
    ("Large curved monitor", "~$1,000", "Premium 34\" QD-OLED ultrawide ~$799 (e.g. Alienware AW3425DW); 49\" super-ultrawide runs higher. (rtings.com, tomsguide.com)"),
    ("Peripherals / desk", "$300", "Keyboard, mouse, cabling, desk odds & ends."),
]
for i,(a,b,c) in enumerate(notes):
    r = 3+i
    ca = n.cell(row=r, column=1, value=a); cb = n.cell(row=r, column=2, value=b); cc = n.cell(row=r, column=3, value=c)
    if i == 0:
        for x in (ca,cb,cc): x.font = hdr_font(); x.fill = head_fill()
    else:
        ca.font = Font(bold=True, color="30425F"); cc.font = Font(color="4A5A72")
        cc.alignment = Alignment(wrap_text=True)
    n.row_dimensions[r].height = 42
n.cell(row=12, column=1,
       value="These are defaults on the Inputs sheet — override any of them with your real bills.").font = Font(italic=True, color="6B7890")

import os
out = os.path.join(os.path.dirname(__file__), "FlipPulse_Business_Model.xlsx")
wb.save(out)
print("wrote", out)
